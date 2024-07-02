import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


def read_excel(file):
    return pd.read_excel(file)


def apply_template_with_formulas_and_validation(raw_data, template_file):
    wb = load_workbook(template_file, data_only=False)
    ws = wb.active

    # Store existing data validations
    original_validations = list(ws.data_validations.dataValidation)

    # Convert raw_data DataFrame to rows
    raw_data_rows = dataframe_to_rows(raw_data, index=False, header=False)

    # Write raw_data to the template sheet, starting from the second row to preserve headers/formulas
    for i, row in enumerate(raw_data_rows, start=2):
        for j, value in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=value)

    # Clear existing data validations
    ws.data_validations.dataValidation.clear()

    # Restore data validations
    for dv in original_validations:
        new_dv = DataValidation(
            type=dv.type,
            operator=dv.operator,
            formula1=dv.formula1,
            formula2=dv.formula2,
            allow_blank=dv.allow_blank,
            showDropDown=dv.showDropDown,
            showInputMessage=dv.showInputMessage,
            showErrorMessage=dv.showErrorMessage,
            errorTitle=dv.errorTitle,
            error=dv.error,
            promptTitle=dv.promptTitle,
            prompt=dv.prompt,
            sqref=dv.sqref,
        )
        ws.add_data_validation(new_dv)

    # Save the updated workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def main():
    st.title("Stage 2: Apply Template to Raw Excel")

    st.markdown(
        """
    ### Instructions:
    1. Upload the Excel template file first.
    2. Then upload a raw Excel file from Stage 1.
    3. Apply the template to the file and download it.
    """
    )

    st.subheader("Upload Template File")
    template_file = st.file_uploader("Upload Template File", type="xlsx")

    if template_file:
        st.write("Template uploaded successfully!")

        st.subheader("Upload Raw Excel File")
        raw_file = st.file_uploader("Upload Raw Excel File", type="xlsx")

        if raw_file:
            raw_data = read_excel(raw_file)
            st.write(f"Raw Excel file '{raw_file.name}' uploaded successfully!")
            st.write(raw_data)

            try:
                applied_data_output = apply_template_with_formulas_and_validation(
                    raw_data, template_file
                )
                st.write("Data with Applied Template:")

                st.download_button(
                    label="Download Excel File with Applied Template",
                    data=applied_data_output,
                    file_name=f"applied_{raw_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            except Exception as e:
                st.error(f"Error applying template: {e}")


if __name__ == "__main__":
    main()
